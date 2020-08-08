import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Border, borders, Side, colors, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, Reference


        # python3 -u "/Users/saurishkapoor/KPMG.py"
def activateFunc():
    excel = pd.ExcelFile("/Users/saurishkapoor/Desktop/KPMG.xlsx")
    df = pd.read_excel(excel,sheet_name=1,)
    df.columns = ["Transaction ID","Product ID","Customer ID","Transaction Date", "Online Order","Order Status", 'Brand','Product Line', 'Product Class', 
    'Product Size', 'List Price', 'Standard Cost', 'First Selling Date']
    df["Index"] = 0
    index_x = df["Index"]

    # Making an index row to sort by
    counts = 0
    for row in range(2, 20003):
        index_x.iloc[row] = counts
        counts = counts + 1

    df.index = df["Index"]

    df.drop(df.index[0], inplace = True)
    df.replace(np.nan, "N/A")
    df["List Price"].replace("N/A",0) 
    df1 = df[["Order Status", "Brand", 'Product Line', 'Product Class', 'Product Size', 'List Price']]
    df1.replace(np.nan, "N/A")
    rep = df1.loc[(df1["Product Line"] == "Standard") & (df1["List Price"] >= 1000)].groupby("Brand").last()
    #print(df1)
    #print(rep)
    rep.to_excel("KPMG_M.xlsx")

    # dtype={"transaction_id": str, "product_id": str, "customer_id": str, 
    # "transaction_date": str , "online_order": str, "order_status": str, "brand": str, "product_line": str, "product_class": str, "product_size": str, 
    #  "list_price": float,"standard_cost": str, "product_first_sold_date": str}

def Format():

    wb = load_workbook('KPMG_M.xlsx')
    ws = wb.active

    ws.merge_cells("A1:F1")
    ws["A1"] = "Analytics and Findings"
    cell = ws["A1"]
    cell.font = Font(bold= True)
    cell.alignment =  Alignment(horizontal= "center")
    side = Side(border_style= "thin", color= colors.BLACK)
    cell.border = Border(bottom= side)
    
    e_col = 'F'
    for row in ws.iter_rows(min_row = 1):
        for c_gix, col in enumerate(row,1):
            ws[e_col +  str(c_gix)].border = Border(right= side)
    
    chart = BarChart()
    chart.type = "col"
    chart.title = "Analytics and Findings"
    data = Reference(min_row = 1, min_col= 1,worksheet= ws, range_string= "Sheet1!A2:F7")
    chart.add_data(data, titles_from_data=True)
    chart.x_axis.title = "amount"
    ws.add_chart(chart, "A10")

    
    wb.save("KPMG_F.xlsx")






        

       








    



Format()










    
